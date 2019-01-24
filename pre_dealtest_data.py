# -*- coding: utf-8 -*-
"""
Created on Tue Nov 14 22:57:29 2017

@author: Administrator
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
"""
wb = Workbook()
ws = wb.active

ws.cell(row=1, column=1).value ='trigger_type'
ws.cell(row=1, column=2).value ='trigger'
ws.cell(row=1, column=3).value ='sentence'
ws.cell(row=1, column=4).value ='end_point'
ws.cell(row=1, column=5).value='sentence_num'
ws.cell(row=1, column=6).value='txt_count'



count=1
hang=2
sentence_count=0
sentence_ccount=0
flag=0
pre_count=2
while count<=259:
        #读取txt（此时的txt是分句后的文档，一行是一句），将每个txt文档的每句提取出来单独分析触发词和类型    
    fp=open('F:\\code_project1\\code2.0\\data_11\\BioNLP-ST_2011_genia_devel_data_rev1\\deal_data\\'+str(count)+'.txt','r')	
    current_txt=fp.read()
    fp.seek(0)
    line_txt=fp.readline()
    while line_txt!='' and sentence_ccount<=len(current_txt.split('\n')):
        if line_txt=='\n': 
            line_txt=fp.readline()
            flag=1
            continue
        ws.cell(row=hang, column=3).value =line_txt
        ws.cell(row=hang, column=5).value = sentence_ccount+1

        if flag==1:
            ws.cell(row=hang, column=4).value =fp.tell()-2
        else:
            ws.cell(row=hang, column=4).value =fp.tell()-ws.cell(row=hang, column=5).value
        ws.cell(row=hang, column=1).value ='null'
        ws.cell(row=hang, column=2).value ='null'
        ws.cell(row=hang, column=6).value = count
        hang=hang+1
        sentence_count=sentence_count+1
        sentence_ccount=sentence_ccount+1
        line_txt=fp.readline()
        
    flag=0
    i=hang-sentence_ccount
    sentence_ccount=0
        #读取txt对应的a2文档，找到触发词和类型
    fp_a2=open('F:\\code_project1\\code2.0\\data_11\\BioNLP-ST_2011_genia_devel_data_rev1\\deal_data\\'+str(count)+'.a2','r')
    
    count=count+1
    
    trigger_results=fp_a2.readlines()
    for line_a2 in trigger_results:
        line_a2_split=line_a2.split('\t')
        if line_a2[0]=='T':
            subsplit=line_a2_split[1].split(' ')
            num_1=int(subsplit[1])
            num_2=int(subsplit[2])
            while i<=hang:
                if (i==pre_count) :
                    if(num_1>ws.cell(row=i,column=4).value):
                        i=i+1
                        continue
                    else:
                        ws.cell(row=i, column=1).value =subsplit[0]
                        ws.cell(row=i, column=2).value =line_a2_split[2][0:len(line_a2_split[2])-1]
                        i=i+1
                        break
                elif ((num_1>ws.cell(row=i-1,column=4).value) and (num_2<ws.cell(row=i,column=4).value)):
                    ws.cell(row=i, column=1).value =subsplit[0]
                    ws.cell(row=i, column=2).value =line_a2_split[2][0:len(line_a2_split[2])-1]
                    i=i+1
                    break
                elif(num_2<ws.cell(row=i-1,column=4).value):
                    sentence=ws.cell(row=i-1,column=3).value
                    ws.cell(row=hang,column=3).value=sentence
                    ws.cell(row=hang, column=1).value =subsplit[0]
                    ws.cell(row=hang,column=2).value=line_a2_split[2][0:len(line_a2_split[2])-1]
                    ws.cell(row=hang,column=4).value=ws.cell(row=i-1,column=4).value 
                    ws.cell(row=hang, column=5).value=ws.cell(row=i-1,column=5).value     
                    ws.cell(row=hang, column=6).value = count-1
                  
                    hang=hang+1
                    sentence_count=sentence_count+1
                    break;
                else:
                    i=i+1
        else:
            i=i+1
    pre_count=hang
wb.save(filename='F:\\code_project1\\code2.0\\data_09\\test_data\\new_testdata\\test2.xlsx')

"""
import xlrd
from openpyxl import Workbook

origin_excel= xlrd.open_workbook('F:\\code_project1\\code2.0\\data_09\\test_data\\new_testdata\\test2.xlsx')
table = origin_excel.sheet_by_name('Sheet')


wb = Workbook()
ws = wb.active

j=0
i=1
while j<4300:
    if(str(table.cell(j,0))!="text:'null'" and str(table.cell(j,0))!="text:'null'"):
        ws.cell(row=i,column=1).value=table.cell(j,0).value
        ws.cell(row=i,column=2).value=table.cell(j,1).value
        ws.cell(row=i,column=3).value=table.cell(j,2).value
        ws.cell(row=i,column=4).value=table.cell(j,3).value
        ws.cell(row=i,column=5).value=table.cell(j,4).value
        ws.cell(row=i,column=6).value=table.cell(j,5).value
        j=j+1
        i=i+1
    else:
        j=j+1

wb.save(filename='F:\\code_project1\\code2.0\\data_09\\test_data\\new_testdata\\test_new3.xlsx')
