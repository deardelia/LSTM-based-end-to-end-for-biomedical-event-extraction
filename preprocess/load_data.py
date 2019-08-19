#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
    Load data.
"""
import pickle
from time import time
import numpy as np
import configurations as config
from TFNN.utils.io_util import read_lines
from TFNN.utils.data_util import map_item2id
import xlrd
from openpyxl import Workbook

def get_sentence_arr(words_tags, word_voc, tag_voc):
    """
    获取词序列
    Args:
        words_tags: list, 句子 and tags
        word_voc: 词表
        tag_voc: 词性标注表
    Returns:
        sentence_arr: np.array, 字符id序列
        tag_arr: np.array, 词性标记序列
    """
    words, postags = [], []
    for item in words_tags:
        if item!='':
            rindex = item.rindex('/')
            words.append(item[:rindex])
            postags.append(item[rindex+1:])
        else:
            continue
    # sentence arr
    sentence_arr = map_item2id(
        words, word_voc, config.MAX_LEN, lower=True)
    # pos tags arr
    postag_arr = map_item2id(
        postags, tag_voc, config.MAX_LEN, lower=False)
    return sentence_arr, postag_arr, len(words)


def init_data(lines, word_voc, tag_voc, label_voc):
    """
    加载数据
    Args:
        lines: list
        word_voc: dict, 词表
        tag_voc: dict, 词性标注表
        label_voc: dict
    Returns:
        sentences: np.array
        etc.
    """
    data_count = len(lines)#表示句子数目
	#sentences矩阵的每行代表一个句子，列数为规定的每句的最大词数（本例中规定为300）
    sentences = np.zeros((data_count, config.MAX_LEN), dtype='int32')
    tags = np.zeros((data_count, config.MAX_LEN), dtype='int32')
	#一维数组分别记录每个句子的实际单词个数
    sentence_actual_lengths = np.zeros((data_count,), dtype='int32')
	#对每个句子有一个标记（本例的标记即为触发词类型
    labels = np.zeros((data_count,), dtype='int32')
    instance_index = 0
    for i in range(data_count):
		#语句和文本的类别格式为“标号,……”
		#lines[i]表示第i个句子
        index = lines[i].index(',')
        label = lines[i][:index]
        sentence = lines[i][index+1:]
        words_tags = sentence.split(' ')
		#sentence_arr代表该句子的单词转换为id后的形式，tag_arr类似，actual_length代表该句子的实际长度（包含的单词数）
        sentence_arr, tag_arr, actual_length = get_sentence_arr(words_tags, word_voc, tag_voc)
		#instance_index代表当前是文本的第几个句子
        sentences[instance_index, :] = sentence_arr
        tags[instance_index, :] = tag_arr
        sentence_actual_lengths[instance_index] = actual_length
        labels[instance_index] = label
        instance_index += 1
    return sentences, tags, labels


def load_embedding():
    """
    加载词向量、词性向量
    Return:
        word_weights: np.array
        tag_weights: np.array
    """
    # 加载词向量
    with open(config.W2V_TRAIN_PATH, 'rb') as file_r:
        word_weights = pickle.load(file_r)
    # 加载tag向量
    with open(config.T2V_PATH, 'rb') as file_r:
        tag_weights = pickle.load(file_r)
    return word_weights, tag_weights


def load_voc():
    """
    Load voc...
    Return:
        word_voc: dict
        tag_voc: dict
        label_voc: dict
    """
    with open(config.WORD_VOC_PATH, 'rb') as file_r:
        word_voc = pickle.load(file_r)
    with open(config.TAG_VOC_PATH, 'rb') as file_r:
        tag_voc = pickle.load(file_r)
    with open(config.LABEL_VOC_PATH, 'rb') as file_r:
        label_voc = pickle.load(file_r)
    return word_voc, tag_voc, label_voc


def load_train_data(word_voc, tag_voc, label_voc,class_type,training_count):
    """
    加载训练测试数据
    Args:
        word_voc: dict
        tag_voc: dict
        label_voc: dict
    Returns:
        xx
    """
    #origin_excel= xlrd.open_workbook("./data_09/training_data/training_dataClass/trainClass_"+str(class_type)+".xlsx")
    origin_excel= xlrd.open_workbook("F:\\code_project1\\code2.0\\data_09\\training_data\\training_dataClass\\trainClass_"+str(class_type)+".xlsx")
    table = origin_excel.sheet_by_name('Sheet')
    lines=[]
    i=1
    while i<training_count:
        sentence=table.cell(i,1).value        
        type_=int(table.cell(i,0).value)
        lines.append(str(type_)+','+sentence)
        i=i+1
	#train_path 是原始数据（即为带词性的语句）
    #return init_data(read_lines(config.TRAIN_PATH), word_voc, tag_voc, label_voc)
    return 	init_data(lines, word_voc, tag_voc, label_voc)


def load_test_data(word_voc, tag_voc, label_voc,class_type,test_count):
    """
    加载测试数据
    Args:
        word_voc: dict
        tag_voc: dict
        label_voc: dict
    Returns:
        xx
    """
    #origin_test_excel= xlrd.open_workbook("./data_09/test_data/training_dataClass/trainClass_"+str(class_type)+".xlsx")
    origin_test_excel= xlrd.open_workbook("F:\\code_project1\\code2.0\\data_09\\test_data\\new_testdata\\testClassnew_"+str(class_type)+".xlsx")
    table = origin_test_excel.sheet_by_name('Sheet')
    lines=[]
    i=1
    while i<test_count:
        sentence=table.cell(i,1).value        
        type_=int(table.cell(i,0).value)
        lines.append(str(type_)+','+sentence)
        i=i+1
    #sentences, tags, _ = init_data(read_lines(config.TEST_PATH), word_voc, tag_voc, label_voc)
    sentences, tags, labels = init_data(lines, word_voc, tag_voc, label_voc)
    return sentences, tags,labels


def demo():
    t0 = time()
    word_weights, tag_weights = load_embedding()
    word_voc, tag_voc, label_voc = load_voc()
    data, label_voc = load_train_data()
    sentences, tags, labels = data[:]
    print(sentences.shape)
    print(tags.shape)
    print(labels.shape)
    print(word_weights.shape)
    print(tag_weights.shape)
    print('Done in %ds!' % (time()-t0))


if __name__ == '__main__':
    demo()
