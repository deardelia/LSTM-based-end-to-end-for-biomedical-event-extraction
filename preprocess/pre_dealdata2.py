# -*- coding: utf-8 -*-
"""
Created on Tue Oct 31 17:26:00 2017
第二步，
目标：根据触发词类型建立对应的触发词库
@author: Administrator
"""
import xlrd
from openpyxl import Workbook

origin_excel= xlrd.open_workbook('F:\\code_project1\\traning_data\\TRAIN\\train_end_new.xlsx')
table = origin_excel.sheet_by_name('Sheet')


wb = Workbook()
ws = wb.active

i=1
j=1

ws.cell(row=1, column=1).value ='type_number'
ws.cell(row=2, column=1).value =1
ws.cell(row=3, column=1).value =2
ws.cell(row=4, column=1).value =3
ws.cell(row=5, column=1).value =4
ws.cell(row=6, column=1).value =5
ws.cell(row=7, column=1).value =6
ws.cell(row=8, column=1).value =7
ws.cell(row=9, column=1).value =8
ws.cell(row=10, column=1).value =9
ws.cell(row=11, column=1).value =10

ws.cell(row=1, column=2).value ='trigger_type'
ws.cell(row=2, column=2).value ='Gene_expression'
ws.cell(row=3, column=2).value ='Transcription'
ws.cell(row=4, column=2).value ='Protein_catabolism'
ws.cell(row=5, column=2).value ='Localization'
ws.cell(row=6, column=2).value ='Binding'
ws.cell(row=7, column=2).value ='Phosphorylation'
ws.cell(row=8, column=2).value ='Regulation'
ws.cell(row=9, column=2).value ='Positive_regulation'
ws.cell(row=10, column=2).value ='Negative_regulation'
ws.cell(row=11, column=2).value ='Entity'

ws.cell(row=1, column=3).value ='trigger_set'

set_1=[]
set_2=[]
set_3=[]
set_4=[]
set_5=[]
set_6=[]
set_7=[]
set_8=[]
set_9=[]
set_10=[]
set_11=[]
while j<=16795:
    trigger_type=table.cell(j,0).value
    if(trigger_type=='Gene_expression'):
        set_1.append(table.cell(j,1).value)
    elif(trigger_type=='Transcription'):
        set_2.append(table.cell(j,1).value)
    elif(trigger_type=='Protein_catabolism'):
        set_3.append(table.cell(j,1).value)
    elif(trigger_type=='Localization'):
        set_4.append(table.cell(j,1).value)
    elif(trigger_type=='Binding'):
        set_5.append(table.cell(j,1).value)
    elif(trigger_type=='Phosphorylation'):
        set_6.append(table.cell(j,1).value)
    elif(trigger_type=='Regulation'):
        set_7.append(table.cell(j,1).value)
    elif(trigger_type=='Positive_regulation'):
        set_8.append(table.cell(j,1).value)
    elif(trigger_type=='Negative_regulation'):
        set_9.append(table.cell(j,1).value)
    elif(trigger_type=='Entity'):
        set_10.append(table.cell(j,1).value)
    else:
        j=j+1
        continue
    j=j+1
set_1=set(set_1)
set_2=set(set_2)
set_3=set(set_3)
set_4=set(set_4)
set_5=set(set_5)
set_6=set(set_6)
set_7=set(set_7)
set_8=set(set_8)
set_9=set(set_9)
set_10=set(set_10)

ws.cell(row=2, column=3).value =','.join(set_1)
ws.cell(row=3, column=3).value =','.join(set_2)
ws.cell(row=4, column=3).value =','.join(set_3)
ws.cell(row=5, column=3).value =','.join(set_4)
ws.cell(row=6, column=3).value =','.join(set_5)
ws.cell(row=7, column=3).value =','.join(set_6)
ws.cell(row=8, column=3).value =','.join(set_7)
ws.cell(row=9, column=3).value =','.join(set_8)
ws.cell(row=10, column=3).value =','.join(set_9)
ws.cell(row=11, column=3).value =','.join(set_10)

wb.save(filename='F:\\code_project1\\traning_data\\TRAIN\\train_trigger.xlsx')




        