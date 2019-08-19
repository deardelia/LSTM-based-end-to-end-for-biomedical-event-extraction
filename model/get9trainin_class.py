# -*- coding: utf-8 -*-
"""
Created on Sun Nov  5 08:41:59 2017

@author: Administrator
"""

import xlrd
from openpyxl import Workbook
from nltk import word_tokenize, pos_tag


origin_excel= xlrd.open_workbook("F:\\code_project1\\code2.0\\data_09\\training_data\\training_dataClass\\train_endTogether.xlsx")
table = origin_excel.sheet_by_name('Sheet')

wb = Workbook()
ws = wb.active

ws.cell(row=1,column=1).value='trigger_class'
ws.cell(row=1,column=2).value='sentence'
ws.cell(row=1,column=3).value='sentence_num'
ws.cell(row=1,column=4).value='txt_num'



count=1
max_count=16796

while count<max_count:
    if(table.cell(count,0).value=='Entity'):
        ws.cell(row=count+1,column=1).value=1
    else:
        ws.cell(row=count+1,column=1).value=0
    sentence=table.cell(count,2).value
    sentence = sentence.replace(',','')
    sentence = sentence.replace('.','')
    sentence = sentence.replace('(','')
    sentence = sentence.replace(')','')
    sentence = sentence.replace('[','')
    sentence = sentence.replace(']','')
    sentence = sentence.replace('\"','')
    sentence = sentence.replace(':','')
    sentence = sentence.replace('\'','')
    sentence = sentence.replace('/',' ')
    sentence = sentence.replace('\n','')
    pos_deal = pos_tag(word_tokenize(sentence))
    new_sentence=''
    i=0
    while i < len(pos_deal):
        new_sentence=new_sentence+pos_deal[i][0]+'/'+pos_deal[i][1]+' '
        i=i+1
    ws.cell(row=count+1,column=2).value=new_sentence
    ws.cell(row=count+1,column=3).value=table.cell(count,4).value
    ws.cell(row=count+1,column=4).value=table.cell(count,5).value
    count=count+1

wb.save(filename='F:\\code_project1\\code2.0\\data_09\\training_data\\training_dataClass\\trainClass_10.xlsx')
